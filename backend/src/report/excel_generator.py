from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from typing import Dict, List, Any, Optional
import json
import os
from datetime import datetime


class ExcelReportGenerator:
    """
    Excel Report Generator using openpyxl
    Creates Excel files based on JSON configuration
    """
    
    def __init__(self):
        self.workbook = None
        self.worksheets = {}
    
    def create_report(self, config: Dict[str, Any], output_path: str) -> str:
        """
        Create Excel report based on JSON configuration
        
        Args:
            config: JSON configuration defining the report structure
            output_path: Path where the Excel file will be saved
            
        Returns:
            str: Path to the created Excel file
        """
        try:
            # Create new workbook
            self.workbook = Workbook()
            
            # Handle sheets configuration
            if 'sheets' not in config or not config['sheets']:
                # Create a default empty sheet if no sheets are configured
                config['sheets'] = [{'name': 'Sheet1'}]
            else:
                # Remove default sheet if we have custom sheets
                self.workbook.remove(self.workbook.active)
            
            # Process each sheet
            for sheet_config in config.get('sheets', []):
                self._create_sheet(sheet_config)
            
            # Apply workbook-level properties
            if 'properties' in config:
                self._apply_workbook_properties(config['properties'])
            
            # Save the workbook
            output_dir = os.path.dirname(output_path)
            if output_dir:  # Only create directory if there is a directory path
                os.makedirs(output_dir, exist_ok=True)
            self.workbook.save(output_path)
            
            return output_path
            
        except Exception as e:
            raise Exception(f"Error creating Excel report: {str(e)}")
    
    def _create_sheet(self, sheet_config: Dict[str, Any]):
        """
        Create a worksheet based on configuration
        """
        sheet_name = sheet_config.get('name', f'Sheet{len(self.worksheets) + 1}')
        
        # Create worksheet
        if len(self.worksheets) == 0 and self.workbook.active:
            ws = self.workbook.active
            ws.title = sheet_name
        else:
            ws = self.workbook.create_sheet(sheet_name)
        
        self.worksheets[sheet_name] = ws
        
        # Apply sheet properties
        if 'properties' in sheet_config:
            self._apply_sheet_properties(ws, sheet_config['properties'])
        
        # Add headers
        if 'headers' in sheet_config:
            self._add_headers(ws, sheet_config['headers'])
        
        # Add data
        if 'data' in sheet_config:
            self._add_data(ws, sheet_config['data'], sheet_config.get('headers', []))
        
        # Apply formatting
        if 'formatting' in sheet_config:
            self._apply_formatting(ws, sheet_config['formatting'])
        
        # Add charts
        if 'charts' in sheet_config:
            self._add_charts(ws, sheet_config['charts'])
        
        # Auto-adjust column widths
        if sheet_config.get('auto_adjust_columns', True):
            self._auto_adjust_columns(ws)
    
    def _apply_workbook_properties(self, properties: Dict[str, Any]):
        """
        Apply workbook-level properties
        """
        if 'title' in properties:
            self.workbook.properties.title = properties['title']
        if 'creator' in properties:
            self.workbook.properties.creator = properties['creator']
        if 'description' in properties:
            self.workbook.properties.description = properties['description']
    
    def _apply_sheet_properties(self, ws, properties: Dict[str, Any]):
        """
        Apply worksheet properties
        """
        if 'tab_color' in properties:
            ws.sheet_properties.tabColor = properties['tab_color']
        if 'zoom' in properties:
            ws.sheet_view.zoomScale = properties['zoom']
    
    def _add_headers(self, ws, headers: List[Dict[str, Any]]):
        """
        Add headers to worksheet
        """
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header.get('title', f'Column {col_idx}')
            
            # Apply header formatting
            if 'style' in header:
                self._apply_cell_style(cell, header['style'])
            else:
                # Default header style
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _add_data(self, ws, data: List[List[Any]], headers: List[Dict[str, Any]]):
        """
        Add data rows to worksheet
        """
        start_row = 2 if headers else 1
        
        for row_idx, row_data in enumerate(data, start_row):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                # Apply column-specific formatting if defined
                if headers and col_idx <= len(headers):
                    header_config = headers[col_idx - 1]
                    if 'data_style' in header_config:
                        self._apply_cell_style(cell, header_config['data_style'])
    
    def _apply_formatting(self, ws, formatting: Dict[str, Any]):
        """
        Apply general formatting to worksheet
        """
        # Apply borders
        if 'borders' in formatting:
            self._apply_borders(ws, formatting['borders'])
        
        # Apply alternating row colors
        if 'alternating_rows' in formatting:
            self._apply_alternating_rows(ws, formatting['alternating_rows'])
        
        # Freeze panes
        if 'freeze_panes' in formatting:
            ws.freeze_panes = formatting['freeze_panes']
    
    def _apply_cell_style(self, cell, style: Dict[str, Any]):
        """
        Apply style to a cell
        """
        if 'font' in style:
            font_config = style['font']
            cell.font = Font(
                name=font_config.get('name', 'Calibri'),
                size=font_config.get('size', 11),
                bold=font_config.get('bold', False),
                italic=font_config.get('italic', False),
                color=font_config.get('color', '000000')
            )
        
        if 'fill' in style:
            fill_config = style['fill']
            cell.fill = PatternFill(
                start_color=fill_config.get('color', 'FFFFFF'),
                end_color=fill_config.get('color', 'FFFFFF'),
                fill_type="solid"
            )
        
        if 'alignment' in style:
            align_config = style['alignment']
            cell.alignment = Alignment(
                horizontal=align_config.get('horizontal', 'general'),
                vertical=align_config.get('vertical', 'bottom'),
                wrap_text=align_config.get('wrap_text', False)
            )
        
        if 'border' in style:
            border_config = style['border']
            side_style = Side(style=border_config.get('style', 'thin'))
            cell.border = Border(
                left=side_style if border_config.get('left', True) else None,
                right=side_style if border_config.get('right', True) else None,
                top=side_style if border_config.get('top', True) else None,
                bottom=side_style if border_config.get('bottom', True) else None
            )
    
    def _apply_borders(self, ws, border_config: Dict[str, Any]):
        """
        Apply borders to a range of cells
        """
        if 'range' in border_config:
            cell_range = border_config['range']
            border_style = Side(style=border_config.get('style', 'thin'))
            border = Border(
                left=border_style,
                right=border_style,
                top=border_style,
                bottom=border_style
            )
            
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = border
    
    def _apply_alternating_rows(self, ws, config: Dict[str, Any]):
        """
        Apply alternating row colors
        """
        start_row = config.get('start_row', 2)
        end_row = config.get('end_row', ws.max_row)
        color1 = config.get('color1', 'FFFFFF')
        color2 = config.get('color2', 'F2F2F2')
        
        for row_idx in range(start_row, end_row + 1):
            color = color1 if row_idx % 2 == 0 else color2
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
    
    def _add_charts(self, ws, charts_config: List[Dict[str, Any]]):
        """
        Add charts to worksheet
        """
        for chart_config in charts_config:
            chart_type = chart_config.get('type', 'bar')
            
            if chart_type == 'bar':
                chart = BarChart()
            elif chart_type == 'line':
                chart = LineChart()
            elif chart_type == 'pie':
                chart = PieChart()
            else:
                continue
            
            # Set chart properties
            chart.title = chart_config.get('title', 'Chart')
            
            # Only set axis titles for charts that have axes (not pie charts)
            if chart_type != 'pie':
                if hasattr(chart, 'x_axis') and chart_config.get('x_axis_title'):
                    chart.x_axis.title = chart_config.get('x_axis_title', 'X Axis')
                if hasattr(chart, 'y_axis') and chart_config.get('y_axis_title'):
                    chart.y_axis.title = chart_config.get('y_axis_title', 'Y Axis')
            
            # Add data to chart
            data_range = chart_config.get('data_range')
            if data_range:
                try:
                    # Handle different data range formats
                    if '!' not in data_range:
                        # Add sheet name if not present
                        data_range = f"{ws.title}!{data_range}"
                    data = Reference(ws, range_string=data_range)
                    chart.add_data(data, titles_from_data=True)
                except Exception as e:
                    print(f"Warning: Could not add chart data range '{data_range}': {e}")
            
            # Set chart position
            position = chart_config.get('position', 'E2')
            ws.add_chart(chart, position)
    
    def _auto_adjust_columns(self, ws):
        """
        Auto-adjust column widths based on content
        """
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width


def generate_excel_report(config_json: str, output_path: str) -> str:
    """
    Convenience function to generate Excel report from JSON string
    
    Args:
        config_json: JSON string containing report configuration
        output_path: Path where Excel file will be saved
        
    Returns:
        str: Path to created Excel file
    """
    try:
        config = json.loads(config_json)
        generator = ExcelReportGenerator()
        return generator.create_report(config, output_path)
    except json.JSONDecodeError as e:
        raise Exception(f"Invalid JSON configuration: {str(e)}")
    except Exception as e:
        raise Exception(f"Error generating report: {str(e)}")


def generate_excel_report_from_dict(config: Dict[str, Any], output_path: str) -> str:
    """
    Convenience function to generate Excel report from dictionary
    
    Args:
        config: Dictionary containing report configuration
        output_path: Path where Excel file will be saved
        
    Returns:
        str: Path to created Excel file
    """
    try:
        generator = ExcelReportGenerator()
        return generator.create_report(config, output_path)
    except Exception as e:
        raise Exception(f"Error generating report: {str(e)}")


# Example usage and testing
if __name__ == "__main__":
    # Example configuration
    example_config = {
        "properties": {
            "title": "Sample Report",
            "creator": "Report Generator",
            "description": "Generated Excel report"
        },
        "sheets": [
            {
                "name": "Sales Data",
                "properties": {
                    "tab_color": "1F4E79",
                    "zoom": 100
                },
                "headers": [
                    {
                        "title": "Product",
                        "style": {
                            "font": {"bold": True, "color": "FFFFFF"},
                            "fill": {"color": "366092"},
                            "alignment": {"horizontal": "center"}
                        }
                    },
                    {
                        "title": "Sales",
                        "style": {
                            "font": {"bold": True, "color": "FFFFFF"},
                            "fill": {"color": "366092"},
                            "alignment": {"horizontal": "center"}
                        }
                    },
                    {
                        "title": "Revenue",
                        "style": {
                            "font": {"bold": True, "color": "FFFFFF"},
                            "fill": {"color": "366092"},
                            "alignment": {"horizontal": "center"}
                        }
                    }
                ],
                "data": [
                    ["Product A", 100, 5000],
                    ["Product B", 150, 7500],
                    ["Product C", 200, 10000]
                ],
                "formatting": {
                    "alternating_rows": {
                        "start_row": 2,
                        "color1": "FFFFFF",
                        "color2": "F2F2F2"
                    },
                    "freeze_panes": "A2"
                },
                "charts": [
                    {
                        "type": "bar",
                        "title": "Sales by Product",
                        "data_range": "A1:C4",
                        "position": "E2",
                        "x_axis_title": "Products",
                        "y_axis_title": "Sales"
                    }
                ]
            }
        ]
    }
    
    # Generate sample report
    output_file = "sample_report.xlsx"
    try:
        result = generate_excel_report_from_dict(example_config, output_file)
        print(f"Report generated successfully: {result}")
    except Exception as e:
        print(f"Error: {e}")