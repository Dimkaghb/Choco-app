from typing import Dict, List, Any, Optional, Union
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle, Color
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.chart import BarChart, LineChart, PieChart
from openpyxl.chart.reference import Reference
import json
import os
from dataclasses import dataclass
from enum import Enum


def normalize_color(color_value: str) -> str:
    """
    Normalize color value by removing # prefix and ensuring proper format
    
    Args:
        color_value: Color string that may contain # prefix
        
    Returns:
        str: Normalized color string without # prefix
    """
    if not color_value:
        return '000000'
    
    # Convert to string and strip whitespace
    color_str = str(color_value).strip()
    
    # Remove # prefix if present
    if color_str.startswith('#'):
        color_str = color_str[1:]
    
    # Ensure we have a valid hex color (6 characters)
    if len(color_str) == 6 and all(c in '0123456789ABCDEFabcdef' for c in color_str):
        return color_str.upper()
    
    # Default to black if invalid
    return '000000'


class CellType(Enum):
    """Enumeration for different cell types"""
    HEADER = "header"
    DATA = "data"
    FORMULA = "formula"
    MERGE = "merge"
    CHART = "chart"
    EMPTY = "empty"


class BorderStyle(Enum):
    """Enumeration for border styles"""
    THIN = "thin"
    MEDIUM = "medium"
    THICK = "thick"
    DOUBLE = "double"
    DOTTED = "dotted"
    DASHED = "dashed"


@dataclass
class CellPosition:
    """Represents a cell position in Excel"""
    row: int
    column: Union[int, str]
    
    def __post_init__(self):
        if isinstance(self.column, str):
            self.column = column_index_from_string(self.column)
    
    @property
    def excel_address(self) -> str:
        """Get Excel address like 'A1'"""
        return f"{get_column_letter(self.column)}{self.row}"
    
    @classmethod
    def from_address(cls, address: str) -> 'CellPosition':
        """Create CellPosition from Excel address like 'A1'"""
        col_str = ''.join(c for c in address if c.isalpha())
        row_str = ''.join(c for c in address if c.isdigit())
        return cls(int(row_str), col_str)


@dataclass
class CellRange:
    """Represents a range of cells in Excel"""
    start: CellPosition
    end: CellPosition
    
    @property
    def excel_range(self) -> str:
        """Get Excel range like 'A1:C3'"""
        return f"{self.start.excel_address}:{self.end.excel_address}"
    
    @classmethod
    def from_range(cls, range_str: str) -> 'CellRange':
        """Create CellRange from Excel range like 'A1:C3'"""
        start_addr, end_addr = range_str.split(':')
        return cls(CellPosition.from_address(start_addr), CellPosition.from_address(end_addr))


@dataclass
class CellStyle:
    """Represents cell styling options"""
    font: Optional[Dict[str, Any]] = None
    fill: Optional[Dict[str, Any]] = None
    border: Optional[Dict[str, Any]] = None
    alignment: Optional[Dict[str, Any]] = None
    number_format: Optional[str] = None
    
    def to_openpyxl_font(self) -> Optional[Font]:
        """Convert to openpyxl Font object"""
        if not self.font:
            return None
        color_value = normalize_color(self.font.get('color', '000000'))
        # Ensure color is in proper aRGB format (8 characters)
        if color_value and len(str(color_value)) == 6:
            color_value = 'FF' + str(color_value)  # Add alpha channel
        color_obj = Color(rgb=color_value) if color_value else None
        return Font(
            name=self.font.get('name', 'Calibri'),
            size=self.font.get('size', 11),
            bold=self.font.get('bold', False),
            italic=self.font.get('italic', False),
            underline=self.font.get('underline', 'none'),
            color=color_obj
        )
    
    def to_openpyxl_fill(self) -> Optional[PatternFill]:
        """Convert to openpyxl PatternFill object"""
        if not self.fill:
            return None
        color_value = normalize_color(self.fill.get('color', 'FFFFFF'))
        return PatternFill(
            start_color=color_value,
            end_color=color_value,
            fill_type=self.fill.get('pattern', 'solid')
        )
    
    def to_openpyxl_border(self) -> Optional[Border]:
        """Convert to openpyxl Border object"""
        if not self.border:
            return None
        
        def create_side(side_config: Dict[str, Any]) -> Side:
            color_value = normalize_color(side_config.get('color', '000000'))
            # Ensure color is in proper aRGB format (8 characters)
            if color_value and len(str(color_value)) == 6:
                color_value = 'FF' + str(color_value)  # Add alpha channel
            color_obj = Color(rgb=color_value) if color_value else None
            return Side(
                border_style=side_config.get('style', 'thin'),
                color=color_obj
            )
        
        return Border(
            left=create_side(self.border.get('left', {})) if self.border.get('left') else None,
            right=create_side(self.border.get('right', {})) if self.border.get('right') else None,
            top=create_side(self.border.get('top', {})) if self.border.get('top') else None,
            bottom=create_side(self.border.get('bottom', {})) if self.border.get('bottom') else None
        )
    
    def to_openpyxl_alignment(self) -> Optional[Alignment]:
        """Convert to openpyxl Alignment object"""
        if not self.alignment:
            return None
        return Alignment(
            horizontal=self.alignment.get('horizontal', 'general'),
            vertical=self.alignment.get('vertical', 'bottom'),
            text_rotation=self.alignment.get('text_rotation', 0),
            wrap_text=self.alignment.get('wrap_text', False),
            shrink_to_fit=self.alignment.get('shrink_to_fit', False),
            indent=self.alignment.get('indent', 0)
        )


@dataclass
class CellDefinition:
    """Represents a complete cell definition"""
    position: CellPosition
    cell_type: CellType
    value: Any = None
    style: Optional[CellStyle] = None
    formula: Optional[str] = None
    merge_range: Optional[CellRange] = None
    comment: Optional[str] = None
    hyperlink: Optional[str] = None
    data_validation: Optional[Dict[str, Any]] = None


@dataclass
class SheetStructure:
    """Represents the structure of an Excel sheet"""
    name: str
    cells: List[CellDefinition]
    column_widths: Optional[Dict[str, float]] = None
    row_heights: Optional[Dict[int, float]] = None
    freeze_panes: Optional[str] = None
    print_area: Optional[str] = None
    page_setup: Optional[Dict[str, Any]] = None
    protection: Optional[Dict[str, Any]] = None


@dataclass
class WorkbookStructure:
    """Represents the structure of an Excel workbook"""
    sheets: List[SheetStructure]
    properties: Optional[Dict[str, Any]] = None
    named_styles: Optional[List[Dict[str, Any]]] = None
    charts: Optional[List[Dict[str, Any]]] = None


class ExcelStructureParser:
    """Parser for Excel file structures using openpyxl"""
    
    def __init__(self):
        self.workbook: Optional[Workbook] = None
        self.structure: Optional[WorkbookStructure] = None
    
    def parse_from_json(self, json_data: Union[str, Dict[str, Any]]) -> WorkbookStructure:
        """
        Parse Excel structure from JSON configuration
        
        Args:
            json_data: JSON string or dictionary containing structure definition
            
        Returns:
            WorkbookStructure object
        """
        if isinstance(json_data, str):
            data = json.loads(json_data)
        else:
            data = json_data
        
        sheets = []
        for sheet_data in data.get('sheets', []):
            cells = []
            
            # Parse cells from sheet data
            for cell_data in sheet_data.get('cells', []):
                position = CellPosition(
                    row=cell_data['position']['row'],
                    column=cell_data['position']['column']
                )
                
                cell_type = CellType(cell_data.get('type', 'data'))
                
                style = None
                if 'style' in cell_data:
                    style = CellStyle(**cell_data['style'])
                
                merge_range = None
                if 'merge_range' in cell_data:
                    merge_data = cell_data['merge_range']
                    merge_range = CellRange(
                        start=CellPosition(merge_data['start']['row'], merge_data['start']['column']),
                        end=CellPosition(merge_data['end']['row'], merge_data['end']['column'])
                    )
                
                cell_def = CellDefinition(
                    position=position,
                    cell_type=cell_type,
                    value=cell_data.get('value'),
                    style=style,
                    formula=cell_data.get('formula'),
                    merge_range=merge_range,
                    comment=cell_data.get('comment'),
                    hyperlink=cell_data.get('hyperlink'),
                    data_validation=cell_data.get('data_validation')
                )
                
                cells.append(cell_def)
            
            sheet_structure = SheetStructure(
                name=sheet_data['name'],
                cells=cells,
                column_widths=sheet_data.get('column_widths'),
                row_heights=sheet_data.get('row_heights'),
                freeze_panes=sheet_data.get('freeze_panes'),
                print_area=sheet_data.get('print_area'),
                page_setup=sheet_data.get('page_setup'),
                protection=sheet_data.get('protection')
            )
            
            sheets.append(sheet_structure)
        
        self.structure = WorkbookStructure(
            sheets=sheets,
            properties=data.get('properties'),
            named_styles=data.get('named_styles'),
            charts=data.get('charts')
        )
        
        return self.structure
    
    def parse_from_file(self, file_path: str) -> WorkbookStructure:
        """
        Parse Excel structure from existing Excel file
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            WorkbookStructure object
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        self.workbook = load_workbook(file_path, data_only=False)
        sheets = []
        
        for sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]
            cells = []
            
            # Parse all cells with content
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None or cell.has_style:
                        position = CellPosition(cell.row, cell.column)
                        
                        # Determine cell type
                        cell_type = CellType.DATA
                        if cell.data_type == 'f':
                            cell_type = CellType.FORMULA
                        elif cell.row == 1:  # Assume first row is header
                            cell_type = CellType.HEADER
                        
                        # Extract style information
                        style = self._extract_cell_style(cell)
                        
                        # Check for merged cells
                        merge_range = None
                        for merged_range in worksheet.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                merge_range = CellRange(
                                    start=CellPosition.from_address(str(merged_range.min_cell)),
                                    end=CellPosition.from_address(str(merged_range.max_cell))
                                )
                                break
                        
                        # Get formula if present
                        formula = None
                        if hasattr(cell, 'data_type') and cell.data_type == 'f':
                            formula = cell.value if cell.value and str(cell.value).startswith('=') else None
                        
                        # Use display value for data cells, formula for formula cells
                        cell_value = cell.displayed_value if hasattr(cell, 'displayed_value') and cell.data_type == 'f' else cell.value
                        
                        cell_def = CellDefinition(
                            position=position,
                            cell_type=cell_type,
                            value=cell_value,
                            style=style,
                            formula=formula,
                            merge_range=merge_range,
                            comment=cell.comment.text if cell.comment else None,
                            hyperlink=cell.hyperlink.target if cell.hyperlink else None
                        )
                        
                        cells.append(cell_def)
            
            # Extract column widths
            column_widths = {}
            for col_letter, col_dimension in worksheet.column_dimensions.items():
                if col_dimension.width:
                    column_widths[col_letter] = col_dimension.width
            
            # Extract row heights
            row_heights = {}
            for row_num, row_dimension in worksheet.row_dimensions.items():
                if row_dimension.height:
                    row_heights[row_num] = row_dimension.height
            
            sheet_structure = SheetStructure(
                name=sheet_name,
                cells=cells,
                column_widths=column_widths if column_widths else None,
                row_heights=row_heights if row_heights else None,
                freeze_panes=worksheet.freeze_panes if worksheet.freeze_panes != 'A1' else None
            )
            
            sheets.append(sheet_structure)
        
        # Extract workbook properties
        properties = {
            'title': self.workbook.properties.title,
            'creator': self.workbook.properties.creator,
            'description': self.workbook.properties.description,
            'subject': self.workbook.properties.subject,
            'keywords': self.workbook.properties.keywords
        }
        
        self.structure = WorkbookStructure(
            sheets=sheets,
            properties=properties
        )
        
        return self.structure
    
    def _extract_cell_style(self, cell) -> Optional[CellStyle]:
        """
        Extract style information from openpyxl cell
        
        Args:
            cell: openpyxl cell object
            
        Returns:
            CellStyle object or None
        """
        if not cell.has_style:
            return None
        
        font_data = None
        if cell.font:
            font_color = None
            if cell.font.color:
                if hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                    color_str = str(cell.font.color.rgb)
                    # Ensure proper format - remove any non-hex characters and ensure 6 or 8 chars
                    color_str = ''.join(c for c in color_str if c.isalnum())
                    if len(color_str) == 8:
                        font_color = color_str[2:]  # Remove alpha channel for storage
                    elif len(color_str) == 6:
                        font_color = color_str
                elif hasattr(cell.font.color, 'value') and cell.font.color.value:
                    color_str = str(cell.font.color.value)
                    color_str = ''.join(c for c in color_str if c.isalnum())
                    if len(color_str) == 8:
                        font_color = color_str[2:]  # Remove alpha channel for storage
                    elif len(color_str) == 6:
                        font_color = color_str
            
            font_data = {
                'name': cell.font.name,
                'size': cell.font.size,
                'bold': cell.font.bold,
                'italic': cell.font.italic,
                'underline': cell.font.underline,
                'color': font_color
            }
        
        fill_data = None
        if cell.fill and cell.fill.start_color:
            fill_data = {
                'color': cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else None,
                'pattern': cell.fill.fill_type
            }
        
        border_data = None
        if cell.border:
            border_data = {}
            for side_name in ['left', 'right', 'top', 'bottom']:
                side = getattr(cell.border, side_name)
                if side and side.style:
                    side_color = None
                    if side.color:
                        if hasattr(side.color, 'rgb') and side.color.rgb:
                            color_str = str(side.color.rgb)
                            color_str = ''.join(c for c in color_str if c.isalnum())
                            if len(color_str) == 8:
                                side_color = color_str[2:]  # Remove alpha channel for storage
                            elif len(color_str) == 6:
                                side_color = color_str
                        elif hasattr(side.color, 'value') and side.color.value:
                            color_str = str(side.color.value)
                            color_str = ''.join(c for c in color_str if c.isalnum())
                            if len(color_str) == 8:
                                side_color = color_str[2:]  # Remove alpha channel for storage
                            elif len(color_str) == 6:
                                side_color = color_str
                    
                    border_data[side_name] = {
                        'style': side.style,
                        'color': side_color
                    }
        
        alignment_data = None
        if cell.alignment:
            alignment_data = {
                'horizontal': cell.alignment.horizontal,
                'vertical': cell.alignment.vertical,
                'text_rotation': cell.alignment.text_rotation,
                'wrap_text': cell.alignment.wrap_text,
                'shrink_to_fit': cell.alignment.shrink_to_fit,
                'indent': cell.alignment.indent
            }
        
        return CellStyle(
            font=font_data,
            fill=fill_data,
            border=border_data,
            alignment=alignment_data,
            number_format=cell.number_format if cell.number_format != 'General' else None
        )
    
    def export_to_json(self, structure: Optional[WorkbookStructure] = None) -> str:
        """
        Export structure to JSON format
        
        Args:
            structure: WorkbookStructure to export (uses self.structure if None)
            
        Returns:
            JSON string representation
        """
        if structure is None:
            structure = self.structure
        
        if structure is None:
            raise ValueError("No structure to export. Parse a structure first.")
        
        data = {
            'properties': structure.properties,
            'sheets': []
        }
        
        for sheet in structure.sheets:
            sheet_data = {
                'name': sheet.name,
                'cells': [],
                'column_widths': sheet.column_widths,
                'row_heights': sheet.row_heights,
                'freeze_panes': sheet.freeze_panes,
                'print_area': sheet.print_area,
                'page_setup': sheet.page_setup,
                'protection': sheet.protection
            }
            
            for cell in sheet.cells:
                cell_data = {
                    'position': {
                        'row': cell.position.row,
                        'column': cell.position.column
                    },
                    'type': cell.cell_type.value,
                    'value': cell.value
                }
                
                if cell.style:
                    cell_data['style'] = {
                        'font': cell.style.font,
                        'fill': cell.style.fill,
                        'border': cell.style.border,
                        'alignment': cell.style.alignment,
                        'number_format': cell.style.number_format
                    }
                
                if cell.formula:
                    cell_data['formula'] = cell.formula
                
                if cell.merge_range:
                    cell_data['merge_range'] = {
                        'start': {
                            'row': cell.merge_range.start.row,
                            'column': cell.merge_range.start.column
                        },
                        'end': {
                            'row': cell.merge_range.end.row,
                            'column': cell.merge_range.end.column
                        }
                    }
                
                if cell.comment:
                    cell_data['comment'] = cell.comment
                
                if cell.hyperlink:
                    cell_data['hyperlink'] = cell.hyperlink
                
                if cell.data_validation:
                    cell_data['data_validation'] = cell.data_validation
                
                sheet_data['cells'].append(cell_data)
            
            data['sheets'].append(sheet_data)
        
        if structure.named_styles:
            data['named_styles'] = structure.named_styles
        
        if structure.charts:
            data['charts'] = structure.charts
        
        return json.dumps(data, indent=2, default=str)
    
    def create_workbook_from_structure(self, structure: Optional[WorkbookStructure] = None) -> Workbook:
        """
        Create openpyxl Workbook from structure
        
        Args:
            structure: WorkbookStructure to create from (uses self.structure if None)
            
        Returns:
            openpyxl Workbook object
        """
        if structure is None:
            structure = self.structure
        
        if structure is None:
            raise ValueError("No structure to create workbook from. Parse a structure first.")
        
        workbook = Workbook()
        
        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # Set workbook properties
        if structure.properties:
            if 'title' in structure.properties:
                workbook.properties.title = structure.properties['title']
            if 'creator' in structure.properties:
                workbook.properties.creator = structure.properties['creator']
            if 'description' in structure.properties:
                workbook.properties.description = structure.properties['description']
        
        # Create sheets
        for sheet_structure in structure.sheets:
            worksheet = workbook.create_sheet(title=sheet_structure.name)
            
            # Apply cells
            for cell_def in sheet_structure.cells:
                cell = worksheet.cell(
                    row=cell_def.position.row,
                    column=cell_def.position.column,
                    value=cell_def.value
                )
                
                # Apply formula if present
                if cell_def.formula:
                    cell.value = cell_def.formula
                
                # Apply styling
                if cell_def.style:
                    if cell_def.style.font:
                        cell.font = cell_def.style.to_openpyxl_font()
                    if cell_def.style.fill:
                        cell.fill = cell_def.style.to_openpyxl_fill()
                    if cell_def.style.border:
                        cell.border = cell_def.style.to_openpyxl_border()
                    if cell_def.style.alignment:
                        cell.alignment = cell_def.style.to_openpyxl_alignment()
                    if cell_def.style.number_format:
                        cell.number_format = cell_def.style.number_format
                
                # Apply comment
                if cell_def.comment:
                    from openpyxl.comments import Comment
                    cell.comment = Comment(cell_def.comment, "System")
                
                # Apply hyperlink
                if cell_def.hyperlink:
                    cell.hyperlink = cell_def.hyperlink
            
            # Apply merged cells
            merged_ranges = set()
            for cell_def in sheet_structure.cells:
                if cell_def.merge_range and cell_def.merge_range.excel_range not in merged_ranges:
                    worksheet.merge_cells(cell_def.merge_range.excel_range)
                    merged_ranges.add(cell_def.merge_range.excel_range)
            
            # Apply column widths
            if sheet_structure.column_widths:
                for col_letter, width in sheet_structure.column_widths.items():
                    worksheet.column_dimensions[col_letter].width = width
            
            # Apply row heights
            if sheet_structure.row_heights:
                for row_num, height in sheet_structure.row_heights.items():
                    worksheet.row_dimensions[row_num].height = height
            
            # Apply freeze panes
            if sheet_structure.freeze_panes:
                worksheet.freeze_panes = sheet_structure.freeze_panes
        
        self.workbook = workbook
        return workbook
    
    def save_workbook(self, file_path: str, structure: Optional[WorkbookStructure] = None) -> None:
        """
        Save workbook to file
        
        Args:
            file_path: Path to save the Excel file
            structure: WorkbookStructure to save (uses self.structure if None)
        """
        workbook = self.create_workbook_from_structure(structure)
        
        # Create directory if it doesn't exist and file_path contains directory
        dir_path = os.path.dirname(file_path)
        if dir_path:  # Only create directory if there is a directory path
            os.makedirs(dir_path, exist_ok=True)
        
        workbook.save(file_path)


def parse_structure(structure_data: Union[str, Dict[str, Any]], file_path: Optional[str] = None) -> WorkbookStructure:
    """
    Main function to parse Excel structure from various sources
    
    Args:
        structure_data: JSON string, dictionary, or file path to existing Excel file
        file_path: Optional output file path for saving
        
    Returns:
        WorkbookStructure object
    """
    parser = ExcelStructureParser()
    
    if isinstance(structure_data, str) and structure_data.endswith(('.xlsx', '.xls')):
        # Parse from existing Excel file
        structure = parser.parse_from_file(structure_data)
    else:
        # Parse from JSON data
        structure = parser.parse_from_json(structure_data)
    
    if file_path:
        parser.save_workbook(file_path, structure)
    
    return structure
